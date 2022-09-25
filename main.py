from datetime import date, timedelta
import datetime
from ics import Calendar,Event
from openpyxl import load_workbook
from chinese_calendar import is_workday


def AddCanlendar(Cal,course):
    #Get the information about week
    week = course[3].split('周')[0]
    IsSingle = False
    IsDouble = False
    if(str(week[len(week) - 1]) == '单'):
        IsSingle = True
    if(str(week[len(week) - 1]) == '双'):
        IsDouble = True
    week = week.replace('单','')
    week = week.replace('双','')
    week_start, week_end = week.split('-')
    #Get the information about location
    location = course[3].split('周')[1].split('-')[0]
    location = location[0:len(location) - 1]
    #Get the information about time
    Time = course[3]
    if(Time[len(Time)-3:len(Time) -1] == '10' or Time[len(Time)-3:len(Time) -1] == '11'):
        Time = Time[len(Time) - 5:len(Time) - 1]
    else:
        Time = Time[len(Time) - 4:len(Time) - 1]
    Time_Start, Time_End = Time.split('-')
    #Translating time
    if(Time_Start == '1'):
        Time_Start ='8:00:00'
    elif(Time_Start == '3'):
        Time_Start = '10:20:00'
    elif(Time_Start == '5'):
        Time_Start = '14:00:00'
    elif(Time_Start == '7'):
        Time_Start = '16:20:00'
    elif(Time_Start == '9'):
        Time_Start = '19:00:00'
    elif(Time_Start == '11'):
        Time_Start = '21:00:00'

    if(Time_End == '2'):
        Time_End ='9:50:00'
    elif(Time_End == '4'):
        Time_End = '12:10:00'
    elif(Time_End == '6'):
        Time_End= '15:50:00'
    elif(Time_End == '8'):
        Time_End = '18:10:00'
    elif(Time_End == '10'):
        Time_End = '20:50:00'
    elif(Time_End == '11'):
        Time_End = '22:00:00'

    #Count Actual Date
    for weeks in range(int(week_start),int(week_end) + 1):
        if(IsDouble == True and weeks%2 == 1):
            continue
        if(IsSingle == True and weeks%2 == 0):
            continue
        Date = date(2022,9,5)
        Date_Bias = (weeks-1)*7 + course[4] -1
        Date = Date + timedelta(days = Date_Bias)
        if(is_workday(Date) is False):
            continue
        #Adding events
        e = Event()
        e.name = course[0]

        e.begin = str(Date) + ' ' + Time_Start
        e.end = str(Date) + ' ' + Time_End
        e.location = location
        Cal.events.add(e)

if __name__ == '__main__':
    #Read the sheet
    wb = load_workbook('学期课表.xlsx')
    sheets = wb.worksheets
    sheet = sheets[0]


    Courses = list()
    for i in range(4, 9):
        for j in range(2, 8):
            cell = sheet.cell(i,j).value
            if(cell is not None):
                #print(cell)
                course_info = cell.split("\n")
                informations = list()
                for info in course_info:
                    informations.append((info.replace('[', '')).replace(']', ''))
                informations.append(j-1)
                Courses.append(informations)
    print(Courses)

    Cal = Calendar()
    for course in Courses:
        AddCanlendar(Cal, course)
    with open('MyCalendar.ics', 'w') as f:
        f.writelines(Cal.serialize_iter())

    #Handle the problem of timezone
    fp = open('MyCalendar.ics', 'r')
    s = fp.read()
    fp.close()
    a = s.split('\n')
    for i in range(0,len(a)):
            a[i] = a[i].replace('DTSTART:','DTSTART;TZID=Asia/Shanghai:')
            a[i] = a[i].replace('DTEND:','DTEND;TZID=Asia/Shanghai:')
    content = (('CALSCALE:GREGORIAN\n') + '\n' +('METHOD:PUBLISH\n')+'\n' +('BEGIN:VTIMEZONE\n')+'\n' +('TZID:Asia/Shanghai\n')+'\n' +('BEGIN:STANDARD\n')+'\n' +('DTSTART;TZID=Asia/Shanghai:19691231T230000\n')+'\n' +('TZOFFSETTO:+100\n')+'\n' +('TZOFFSETFROM:+100\n')+'\n' +('END:STANDARD\n')+'\n' +('END:VTIMEZONE\n'))
    a.insert(6, content)
    s = '\n'.join(a)
    fp = open('MyCalendar.ics', 'w')
    fp.write(s)
    fp.close()





